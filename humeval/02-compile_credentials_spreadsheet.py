# %%

import glob
import csv
import openpyxl
import collections
import openpyxl.cell

data_all = collections.defaultdict(list)
# for f_name in sorted(glob.glob("/home/vilda/Downloads/wmt25.wave1_v5.accounts/accounts/*.csv")):
# for f_name in sorted(glob.glob("/home/vilda/Downloads/wmt25.wave1_v5ma.accounts/accounts/*.csv")):
# for f_name in sorted(glob.glob("/home/vilda/Downloads/wmt25.wave2_v6.accounts/accounts/*.csv")):
for f_name in sorted(glob.glob("/home/vilda/Downloads/wmt25.ctrl_v7.accounts/accounts/*.csv")):
    langs, domain, wave = (
        f_name.split("/")[-1]
        .removeprefix("wmt25")
        .removesuffix("v5")
        .removesuffix(".csv")
        .split("I")
    )
    langs = langs[:3] + "-" + langs[3:]

    with open(f_name, "r") as f:
        data = [(x["URL"], x["ConfirmationToken"]) for x in csv.DictReader(f)]

    data_all[langs].append((domain, wave, data))

# %%
# create new worksheet
wb = openpyxl.Workbook()
# add new worksheet (not ws.active)

for langs, data_multi in data_all.items():
    ws = wb.create_sheet(langs)

    campaign_keys = [
        f"wmt25{langs.replace("-", "")}I{domain}I{wave}"
        for domain, wave, _ in data_multi
    ]
    link_status = f"https://appraise-wmt.azurewebsites.net/campaign-status/{",".join(campaign_keys)}/"
    ws.append(["Monitor", link_status])
    for cell in ws[f"A1:B1"][0]:
        cell.fill = openpyxl.styles.PatternFill(start_color="edaa6f", end_color="edaa6f", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True)
    
    ws.append(["URL", "Token", "Domain", "Wave", "User"])
    for cell in ws["A2:E2"][0]:
        cell.font = openpyxl.styles.Font(bold=True)

    
    for domain, wave, data in data_multi:
        ws.append([])
        for i, (url, token) in enumerate(data):
            ws.append([url, token, domain, wave])

# wb.save("/home/vilda/Downloads/credentials_wave1v5.xlsx")
wb.save("/home/vilda/Downloads/credentials_wave1v7.xlsx")